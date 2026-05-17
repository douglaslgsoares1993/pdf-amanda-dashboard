"""Banco SQLite e artefatos da camada estruturada."""

from __future__ import annotations

import html
import json
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
SAIDA = BASE_DIR / "SAIDA"
DB_PATH = SAIDA / "procedimentos.db"
XLSX_ESTRUTURADO = SAIDA / "DADOS_ESTRUTURADOS.xlsx"
PESQUISA_XLSX = SAIDA / "PESQUISA_CLINICA.xlsx"
DASHBOARD_HTML = SAIDA / "DASHBOARD.html"

COR_AZUL_ESCURO = "1F3864"
COR_AZUL_MEDIO = "2E74B5"
COR_AZUL_CLARO = "D6E4F0"
COR_LINHA_ALTERNADA = "F7FBFF"
COR_BRANCO = "FFFFFF"
COR_BORDA = "BFBFBF"
COR_AMARELO_USUARIO = "FFFDE7"
COR_AZUL_SISTEMA = "E3F2FD"
COR_TOTAL = "FFF2CC"

CAMPOS = [
    "prontuario", "nome", "cpf", "dt_nascimento", "idade", "sexo", "municipio", "servico",
    "unidade_internacao", "leito", "tipo_atendimento", "origem_paciente", "cod_atendimento",
    "data_atendimento", "data_alta", "cirurgia", "robotica", "lateralidade", "data_inicio",
    "data_fim", "duracao_minutos", "cirurgiao", "especialidade", "anestesista", "centro_sala",
    "arquivo_origem", "tipo_procedimento",
]


def conectar() -> sqlite3.Connection:
    SAIDA.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def inicializar_banco(reprocessar: bool = False) -> None:
    with conectar() as conn:
        if reprocessar:
            conn.execute("DROP TABLE IF EXISTS registros_estruturados")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS registros_estruturados (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                prontuario TEXT, nome TEXT, cpf TEXT, dt_nascimento TEXT, idade INTEGER, sexo TEXT,
                municipio TEXT, servico TEXT, unidade_internacao TEXT, leito TEXT,
                tipo_atendimento TEXT, origem_paciente TEXT, cod_atendimento TEXT,
                data_atendimento TEXT, data_alta TEXT, cirurgia TEXT, robotica TEXT,
                lateralidade TEXT, data_inicio TEXT, data_fim TEXT, duracao_minutos INTEGER,
                cirurgiao TEXT, especialidade TEXT, anestesista TEXT, centro_sala TEXT,
                arquivo_origem TEXT, tipo_procedimento TEXT
            )
            """
        )
        for idx in ("nome", "cpf", "prontuario", "cirurgiao", "data_inicio", "tipo_procedimento"):
            conn.execute(f"CREATE INDEX IF NOT EXISTS idx_reg_{idx} ON registros_estruturados({idx})")


def inserir_registros(registros: list[dict[str, Any]]) -> int:
    if not registros:
        return 0
    inicializar_banco(False)
    sql = f"INSERT INTO registros_estruturados ({','.join(CAMPOS)}) VALUES ({','.join('?' for _ in CAMPOS)})"
    valores = [[r.get(c) for c in CAMPOS] for r in registros]
    fontes = sorted({(r.get("arquivo_origem") or "", r.get("tipo_procedimento") or "") for r in registros})
    with conectar() as conn:
        for arquivo_origem, tipo_procedimento in fontes:
            conn.execute(
                "DELETE FROM registros_estruturados WHERE arquivo_origem = ? AND tipo_procedimento = ?",
                (arquivo_origem, tipo_procedimento),
            )
        antes = conn.total_changes
        conn.executemany(sql, valores)
        return conn.total_changes - antes


def listar_registros() -> list[dict[str, Any]]:
    if not DB_PATH.exists():
        return []
    with conectar() as conn:
        rows = conn.execute("SELECT * FROM registros_estruturados ORDER BY data_inicio DESC, nome").fetchall()
    return [dict(r) for r in rows]


def total_banco() -> int:
    if not DB_PATH.exists():
        return 0
    with conectar() as conn:
        return int(conn.execute("SELECT COUNT(*) FROM registros_estruturados").fetchone()[0])


def buscar_registros(**filtros: Any) -> list[dict[str, Any]]:
    if not DB_PATH.exists():
        return []
    where: list[str] = []
    params: list[Any] = []
    like_fields = ["nome", "cpf", "prontuario", "cirurgiao", "municipio"]
    for field in like_fields:
        valor = filtros.get(field)
        if valor:
            where.append(f"{field} LIKE ?")
            params.append(f"%{valor}%")
    data_ini = filtros.get("data_inicio")
    data_fim = filtros.get("data_fim")
    if data_ini:
        where.append("substr(data_inicio, 7, 4) || '-' || substr(data_inicio, 4, 2) || '-' || substr(data_inicio, 1, 2) >= ?")
        params.append(data_ini)
    if data_fim:
        where.append("substr(data_inicio, 7, 4) || '-' || substr(data_inicio, 4, 2) || '-' || substr(data_inicio, 1, 2) <= ?")
        params.append(data_fim)
    sql = "SELECT * FROM registros_estruturados"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY data_inicio DESC, nome LIMIT 1000"
    with conectar() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def completar_pesquisa(prontuario: int | str) -> list[dict[str, Any]]:
    regs = buscar_registros(prontuario=str(prontuario))
    return [_mapear_pesquisa(r) for r in regs]


def _lado(valor: str) -> str:
    v = (valor or "").upper()
    if "DIREITA" in v:
        return "D"
    if "ESQUERDA" in v:
        return "E"
    if "BILATERAL" in v:
        return "B"
    if "NÃO SE APLICA" in v or "NAO SE APLICA" in v:
        return "NA"
    return valor or ""


def _data(valor: str) -> str:
    return (valor or "")[:10]


def _mapear_pesquisa(r: dict[str, Any]) -> dict[str, Any]:
    return {
        "Data do Procedimento": _data(r.get("data_inicio", "")),
        "Idade": r.get("idade", ""),
        "DN": r.get("dt_nascimento", ""),
        "Sexo": r.get("sexo", ""),
        "LADO": _lado(r.get("lateralidade", "")),
        "Nome Completo": r.get("nome", ""),
        "Tipo de Procedimento": r.get("tipo_procedimento", ""),
        "Cirurgião": r.get("cirurgiao", ""),
        "Origem": r.get("origem_paciente", ""),
        "Município": r.get("municipio", ""),
    }


def _border() -> Border:
    side = Side(style="thin", color=COR_BORDA)
    return Border(left=side, right=side, top=side, bottom=side)


def _header(ws) -> None:
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=COR_AZUL_ESCURO)
        cell.font = Font(bold=True, color=COR_BRANCO)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.zoomScale = 90


def _style_data(ws, start_row: int = 2) -> None:
    for row in ws.iter_rows(min_row=start_row):
        fill = PatternFill("solid", fgColor=COR_LINHA_ALTERNADA if row[0].row % 2 == 0 else COR_BRANCO)
        for cell in row:
            cell.border = _border()
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _ajustar(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        letra = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[letra].width = min(max(max_len + 4, 12), max_width)


def _parse_dt(valor: str) -> datetime | None:
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(valor or "", fmt)
        except ValueError:
            pass
    return None


def _dias(data1: str, data2: str | None) -> int | None:
    d1 = _parse_dt(data1)
    d2 = _parse_dt(data2 or "")
    if not d1 or not d2:
        return None
    return max(0, (d2.date() - d1.date()).days)


def gerar_excel_estruturado() -> Path:
    registros = listar_registros()
    wb = Workbook()
    ws = wb.active
    ws.title = "Todos os Procedimentos"
    ws.append([
        "Tipo de Procedimento", "Prontuário", "Nome do Paciente", "CPF", "Data de Nascimento", "Idade", "Sexo",
        "Município", "Tipo de Atendimento", "Origem do Paciente", "Código de Atendimento", "Data de Internação",
        "Data de Alta", "Dias Internado", "Cirurgia (nome completo)", "Lateralidade", "Data de Início do Procedimento",
        "Data de Fim do Procedimento", "Duração (minutos)", "Cirurgião", "Anestesista", "Centro / Sala",
        "Serviço", "Unidade de Internação", "Leito", "Robótica", "Arquivo de Origem",
    ])
    for r in registros:
        ws.append([
            r["tipo_procedimento"], r["prontuario"], r["nome"], r["cpf"], r["dt_nascimento"], r["idade"], r["sexo"],
            r["municipio"], r["tipo_atendimento"], r["origem_paciente"], r["cod_atendimento"], r["data_atendimento"],
            r["data_alta"], _dias(r["data_atendimento"], r["data_alta"]), r["cirurgia"], r["lateralidade"],
            r["data_inicio"], r["data_fim"], r["duracao_minutos"], r["cirurgiao"], r["anestesista"], r["centro_sala"],
            r["servico"], r["unidade_internacao"], r["leito"], r["robotica"], r["arquivo_origem"],
        ])
    _header(ws)
    _style_data(ws)

    _aba_resumo_tipo(wb, registros)
    _aba_ranking_cirurgioes(wb, registros)
    _aba_multiplos(wb, registros)
    _aba_distribuicao(wb, registros, "Por Município", "municipio")
    _aba_distribuicao(wb, registros, "Lateralidade", "lateralidade")
    _aba_distribuicao(wb, registros, "Origem do Paciente", "origem_paciente")
    busca = wb.create_sheet("Busca")
    busca.append(["Termo", "Nome", "CPF", "Prontuário", "Procedimento", "Data", "Cirurgião", "Arquivo", "Data da Busca"])
    _header(busca)

    for sheet in wb.worksheets:
        _ajustar(sheet)
        sheet.sheet_view.zoomScale = 90
    wb.active = 0
    wb.save(XLSX_ESTRUTURADO)
    return XLSX_ESTRUTURADO


def _aba_resumo_tipo(wb: Workbook, registros: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Resumo por Tipo de Procedimento")
    ws.append(["Tipo", "Total", "Período", "Pacientes únicos", "Cirurgiões distintos", "Duração média", "Duração mín", "Duração máx", "Com anestesista", "Sem anestesista", "Ambulatorial", "Internação", "Com alta", "Sem alta"])
    for tipo in sorted({r["tipo_procedimento"] for r in registros}):
        regs = [r for r in registros if r["tipo_procedimento"] == tipo]
        datas = sorted(_data(r["data_inicio"]) for r in regs if r.get("data_inicio"))
        dur = [r["duracao_minutos"] for r in regs if r.get("duracao_minutos") is not None]
        ws.append([
            tipo, len(regs), f"{datas[0]} a {datas[-1]}" if datas else "", len({r["cpf"] for r in regs if r["cpf"]}),
            len({r["cirurgiao"] for r in regs if r["cirurgiao"]}), round(sum(dur) / len(dur), 1) if dur else "",
            min(dur) if dur else "", max(dur) if dur else "", sum(1 for r in regs if r["anestesista"]),
            sum(1 for r in regs if not r["anestesista"]), sum(1 for r in regs if "Ambulatorial" in r["tipo_atendimento"]),
            sum(1 for r in regs if "Internação" in r["tipo_atendimento"]), sum(1 for r in regs if r["data_alta"]),
            sum(1 for r in regs if not r["data_alta"]),
        ])
    _header(ws)
    _style_data(ws)


def _aba_ranking_cirurgioes(wb: Workbook, registros: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Ranking de Cirurgiões")
    tipos = sorted({r["tipo_procedimento"] for r in registros})
    ws.append(["Cirurgião", "Total", *tipos, "Duração média", "Primeiro procedimento", "Último procedimento", "Pacientes únicos"])
    for cir, regs in sorted(_grupo(registros, "cirurgiao").items(), key=lambda kv: len(kv[1]), reverse=True):
        dur = [r["duracao_minutos"] for r in regs if r.get("duracao_minutos") is not None]
        datas = sorted(_data(r["data_inicio"]) for r in regs if r.get("data_inicio"))
        ws.append([cir, len(regs), *[sum(1 for r in regs if r["tipo_procedimento"] == t) for t in tipos], round(sum(dur) / len(dur), 1) if dur else "", datas[0] if datas else "", datas[-1] if datas else "", len({r["cpf"] for r in regs if r["cpf"]})])
    _header(ws)
    _style_data(ws)


def _aba_multiplos(wb: Workbook, registros: list[dict[str, Any]]) -> None:
    # Excel limita nomes de abas a 31 caracteres.
    ws = wb.create_sheet("Pacientes com Múltiplos Proc")
    ws.append(["Nome", "CPF", "Dt Nascimento", "Município", "Total", "Tipos realizados", "Datas", "Cirurgiões"])
    for cpf, regs in sorted(_grupo(registros, "cpf").items(), key=lambda kv: len(kv[1]), reverse=True):
        if not cpf or len(regs) <= 1:
            continue
        r0 = regs[0]
        ws.append([r0["nome"], cpf, r0["dt_nascimento"], r0["municipio"], len(regs), " | ".join(sorted({r["tipo_procedimento"] for r in regs})), " | ".join(_data(r["data_inicio"]) for r in regs), " | ".join(sorted({r["cirurgiao"] for r in regs if r["cirurgiao"]}))])
    _header(ws)
    _style_data(ws)


def _aba_distribuicao(wb: Workbook, registros: list[dict[str, Any]], nome_aba: str, campo: str) -> None:
    ws = wb.create_sheet(nome_aba)
    tipos = sorted({r["tipo_procedimento"] for r in registros})
    ws.append([nome_aba, "Total", *tipos])
    for valor, regs in sorted(_grupo(registros, campo).items(), key=lambda kv: len(kv[1]), reverse=True):
        ws.append([valor or "(vazio)", len(regs), *[sum(1 for r in regs if r["tipo_procedimento"] == t) for t in tipos]])
    _header(ws)
    _style_data(ws)


def _grupo(registros: list[dict[str, Any]], campo: str) -> dict[str, list[dict[str, Any]]]:
    grupos: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in registros:
        grupos[r.get(campo) or ""].append(r)
    return grupos


def gerar_pesquisa_clinica(prontuarios: list[int | str] | None = None, caminho: Path = PESQUISA_XLSX) -> tuple[Path, int, list[str]]:
    if prontuarios is None:
        prontuarios = sorted({r["prontuario"] for r in listar_registros()})
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados da Pesquisa"
    ws.merge_cells("A1:N1")
    ws["A1"] = "Colunas amarelas: preencher manualmente | Colunas azuis: preenchidas automaticamente pelo sistema"
    ws["A1"].font = Font(bold=True, color=COR_AZUL_ESCURO)
    headers = ["ID", "Pront", "Data do Procedimento", "Idade", "DN", "Sexo", "LADO", "IMC", "Tab", "Nome Completo", "Tipo de Procedimento", "Cirurgião", "Origem", "Município"]
    ws.append(headers)
    encontrados = 0
    nao = []
    row = 3
    for pront in prontuarios:
        dados = completar_pesquisa(pront)
        if not dados:
            ws.append(["", str(pront), "", "", "", "", "", "", "", "", "", "", "", ""])
            nao.append(str(pront))
            row += 1
            continue
        for item in dados:
            ws.append(["", str(pront), item["Data do Procedimento"], item["Idade"], item["DN"], item["Sexo"], item["LADO"], "", "", item["Nome Completo"], item["Tipo de Procedimento"], item["Cirurgião"], item["Origem"], item["Município"]])
            encontrados += 1
            row += 1
    for cell in ws[2]:
        cell.font = Font(bold=True, color=COR_BRANCO)
        cell.fill = PatternFill("solid", fgColor=COR_AZUL_ESCURO)
        cell.border = _border()
    for row_cells in ws.iter_rows(min_row=3):
        for idx, cell in enumerate(row_cells, start=1):
            cell.fill = PatternFill("solid", fgColor=COR_AMARELO_USUARIO if idx in (1, 2, 8, 9) else COR_AZUL_SISTEMA)
            cell.border = _border()
    ws.column_dimensions["J"].width = 0
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:N{ws.max_row}"
    _ajustar(ws)
    ws.column_dimensions["J"].width = 0

    imp = wb.create_sheet("Importar da Planilha Existente")
    imp["A1"] = "Cole aqui sua planilha existente (prontuários e IDs que você já registrou). O sistema vai completar automaticamente os campos do banco de dados."
    imp["B1"].comment = Comment("Após colar, vá ao menu do sistema e escolha opção [7] Atualizar Planilha de Pesquisa", "PDF Amanda System")
    _ajustar(imp)
    wb.save(caminho)
    return caminho, encontrados, nao


def completar_planilha_existente(caminho: Path) -> tuple[int, list[str]]:
    wb = load_workbook(caminho)
    ws = wb["Dados da Pesquisa"] if "Dados da Pesquisa" in wb.sheetnames else wb.active
    headers = {str(c.value).strip(): c.column for c in ws[2] if c.value}
    if "Pront" not in headers:
        headers = {str(c.value).strip(): c.column for c in ws[1] if c.value}
        start_row = 2
    else:
        start_row = 3
    completados = 0
    nao: list[str] = []
    for row in range(start_row, ws.max_row + 1):
        pront = ws.cell(row, headers.get("Pront", 2)).value
        if not pront:
            continue
        dados = completar_pesquisa(str(pront))
        if not dados:
            nao.append(str(pront))
            continue
        item = dados[0]
        for campo in ["Data do Procedimento", "Idade", "DN", "Sexo", "LADO", "Nome Completo", "Tipo de Procedimento", "Cirurgião", "Origem", "Município"]:
            if campo in headers:
                ws.cell(row, headers[campo]).value = item.get(campo, "")
        completados += 1
    wb.save(caminho)
    return completados, nao


def gerar_dashboard_html() -> Path:
    registros = listar_registros()
    gerado = datetime.now().strftime("%d/%m/%Y %H:%M")
    dados = {"registros": registros, "gerado_em": gerado}
    html_text = _html_dashboard(dados)
    DASHBOARD_HTML.write_text(html_text, encoding="utf-8")
    return DASHBOARD_HTML


def _html_dashboard(dados: dict[str, Any]) -> str:
    payload = json.dumps(dados, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="pt-br">
<head>
<meta charset="utf-8">
<title>Dashboard PDF Amanda</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
body {{ font-family: Arial, sans-serif; margin: 0; background: #f5f7fb; color: #1f2937; }}
header {{ background: #1F3864; color: white; padding: 24px 32px; }}
main {{ padding: 24px; }}
.cards {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; }}
.card, section {{ background: white; border: 1px solid #d8dee9; border-radius: 8px; padding: 16px; margin-bottom: 18px; }}
.card b {{ display: block; font-size: 28px; color: #1F3864; }}
.grid2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }}
canvas {{ max-height: 320px; }}
input {{ padding: 10px; width: 360px; max-width: 100%; }}
button {{ padding: 8px 12px; margin: 4px; border: 1px solid #1F3864; background: white; color: #1F3864; border-radius: 5px; cursor: pointer; }}
button.active {{ background: #1F3864; color: white; }}
table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
th {{ background: #1F3864; color: white; cursor: pointer; }}
td, th {{ border: 1px solid #BFBFBF; padding: 6px; text-align: left; }}
tr:nth-child(even) {{ background: #F7FBFF; }}
.badge {{ border-radius: 10px; padding: 3px 7px; color: white; font-weight: bold; }}
.D {{ background: #2E74B5; }} .E {{ background: #2e7d32; }} .NA {{ background: #777; }} .B {{ background: #7b1fa2; }}
</style>
</head>
<body>
<header><h1>Hospital Universitário Pedro Ernesto / UERJ</h1><p>Dashboard gerado em {html.escape(dados['gerado_em'])}</p></header>
<main>
<div class="cards">
<div class="card">Total de procedimentos<b id="m_total">0</b></div>
<div class="card">Pacientes únicos<b id="m_pac">0</b></div>
<div class="card">Cirurgiões distintos<b id="m_cir">0</b></div>
<div class="card">Duração média<b id="m_dur">0 min</b></div>
</div>
<div class="grid2"><section><h2>Total por tipo</h2><canvas id="tipo"></canvas></section><section><h2>Lateralidade</h2><canvas id="lat"></canvas></section></div>
<section><h2>Evolução mensal</h2><canvas id="mes"></canvas></section>
<div class="grid2"><section><h2>Top 10 cirurgiões</h2><canvas id="cir"></canvas></section><section><h2>Internação vs ambulatorial</h2><canvas id="atend"></canvas></section></div>
<section><h2>Top 10 municípios</h2><canvas id="mun"></canvas></section>
<section><h2>Busca interativa</h2><input id="q" placeholder="Nome, prontuário, CPF, cirurgião, município"><div id="tabs"></div><p id="count"></p><table><thead><tr><th>Prontuário</th><th>Nome</th><th>Procedimento</th><th>Data</th><th>Lateralidade</th><th>Cirurgião</th><th>Município</th></tr></thead><tbody id="tbody"></tbody></table><button onclick="page-- ; renderTable()">Anterior</button><button onclick="page++ ; renderTable()">Próxima</button></section>
<section><h2>Cirurgiões</h2><table id="tcir"></table></section>
<section><h2>Pacientes com múltiplos procedimentos</h2><table id="tmulti"></table></section>
</main>
<script>
const DADOS = {payload};
const R = DADOS.registros;
let filtro='Todos', page=0, sortKey='';
const tipos = ['Todos', ...Array.from(new Set(R.map(r=>r.tipo_procedimento))).sort()];
document.getElementById('tabs').innerHTML = tipos.map(t=>`<button onclick="filtro='${{t}}';page=0;renderTable()" id="tab_${{t}}">${{t}}</button>`).join('');
function countBy(arr, key) {{ const o={{}}; arr.forEach(r=>{{const k=typeof key==='function'?key(r):(r[key]||''); o[k]=(o[k]||0)+1}}); return o; }}
function top(obj,n=10) {{ return Object.entries(obj).sort((a,b)=>b[1]-a[1]).slice(0,n); }}
function lado(v) {{ v=(v||'').toUpperCase(); if(v.includes('DIREITA'))return'D'; if(v.includes('ESQUERDA'))return'E'; if(v.includes('BILATERAL'))return'B'; return'NA'; }}
document.getElementById('m_total').textContent=R.length;
document.getElementById('m_pac').textContent=new Set(R.map(r=>r.cpf).filter(Boolean)).size;
document.getElementById('m_cir').textContent=new Set(R.map(r=>r.cirurgiao).filter(Boolean)).size;
const dur=R.map(r=>r.duracao_minutos).filter(x=>x!==null); document.getElementById('m_dur').textContent=Math.round(dur.reduce((a,b)=>a+b,0)/dur.length)+' min';
function chart(id,type,labels,data,opts={{}}) {{ new Chart(document.getElementById(id), {{type, data: {{labels, datasets:[{{label:'Total', data, backgroundColor:['#1F3864','#2E74B5','#7BA7D7','#9CCC65','#FFB74D','#9575CD','#4DB6AC','#E57373','#A1887F','#90A4AE']}}]}}, options: {{responsive:true, ...opts}} }}); }}
let c=countBy(R,'tipo_procedimento'); chart('tipo','bar',Object.keys(c),Object.values(c));
c=countBy(R,r=>lado(r.lateralidade)); chart('lat','doughnut',Object.keys(c),Object.values(c));
c=countBy(R,r=>(r.data_inicio||'').slice(3,10)); chart('mes','line',Object.keys(c).sort(),Object.keys(c).sort().map(k=>c[k]));
c=Object.fromEntries(top(countBy(R,'cirurgiao'))); chart('cir','bar',Object.keys(c),Object.values(c),{{indexAxis:'y'}});
c=countBy(R,'tipo_atendimento'); chart('atend','doughnut',Object.keys(c),Object.values(c));
c=Object.fromEntries(top(countBy(R,'municipio'))); chart('mun','bar',Object.keys(c),Object.values(c));
function filtrados() {{ let q=document.getElementById('q').value.toLowerCase(); return R.filter(r=>(filtro==='Todos'||r.tipo_procedimento===filtro) && [r.nome,r.prontuario,r.cpf,r.cirurgiao,r.municipio].join(' ').toLowerCase().includes(q)); }}
function renderTable() {{ document.querySelectorAll('#tabs button').forEach(b=>b.classList.remove('active')); let tb=document.getElementById('tab_'+filtro); if(tb)tb.classList.add('active'); let a=filtrados(); let pages=Math.max(1,Math.ceil(a.length/50)); page=Math.max(0,Math.min(page,pages-1)); document.getElementById('count').textContent=`${{a.length}} resultado(s)`; document.getElementById('tbody').innerHTML=a.slice(page*50,page*50+50).map(r=>`<tr><td>${{r.prontuario}}</td><td>${{r.nome}}</td><td>${{r.tipo_procedimento}}</td><td>${{(r.data_inicio||'').slice(0,10)}}</td><td><span class="badge ${{lado(r.lateralidade)}}">${{lado(r.lateralidade)}}</span></td><td>${{r.cirurgiao}}</td><td>${{r.municipio}}</td></tr>`).join(''); }}
document.getElementById('q').addEventListener('input',()=>{{page=0;renderTable()}});
function table(id, rows, heads) {{ document.getElementById(id).innerHTML='<thead><tr>'+heads.map(h=>`<th>${{h}}</th>`).join('')+'</tr></thead><tbody>'+rows.map(r=>'<tr>'+r.map(c=>`<td>${{c}}</td>`).join('')+'</tr>').join('')+'</tbody>'; }}
const byCir=countBy(R,'cirurgiao'); table('tcir', top(byCir, 999).map(([cir,total])=>{{let regs=R.filter(r=>r.cirurgiao===cir); return [cir,total,...['Arteriografia','Aortografia','Angioplastia'].map(t=>regs.filter(r=>r.tipo_procedimento===t).length), Math.round(regs.reduce((a,b)=>a+(b.duracao_minutos||0),0)/regs.length), regs.map(r=>r.data_inicio).sort()[0], regs.map(r=>r.data_inicio).sort().at(-1)];}}), ['Cirurgião','Total','Arteriografia','Aortografia','Angioplastia','Duração média','Primeiro','Último']);
const byCpf={{}}; R.forEach(r=>{{if(r.cpf)(byCpf[r.cpf] ||= []).push(r)}}); table('tmulti', Object.values(byCpf).filter(x=>x.length>1).sort((a,b)=>b.length-a.length).map(regs=>[regs[0].nome,regs[0].cpf,regs.length,[...new Set(regs.map(r=>r.tipo_procedimento))].join(' | '),regs.map(r=>r.data_inicio.slice(0,10)).join(' | ')]), ['Nome','CPF','Total','Tipos','Datas']);
renderTable();
</script></body></html>"""


def imprimir_tabela_terminal(registros: list[dict[str, Any]], limite: int = 30) -> str:
    headers = ["Prontuário", "Nome", "Procedimento", "Data", "Cirurgião", "Município"]
    rows = [[r["prontuario"], r["nome"][:28], r["tipo_procedimento"], _data(r["data_inicio"]), r["cirurgiao"][:28], r["municipio"][:22]] for r in registros[:limite]]
    widths = [max(len(str(x)) for x in [h] + [row[i] for row in rows]) for i, h in enumerate(headers)]
    out = [" | ".join(h.ljust(widths[i]) for i, h in enumerate(headers))]
    out.append("-+-".join("-" * w for w in widths))
    out.extend(" | ".join(str(v).ljust(widths[i]) for i, v in enumerate(row)) for row in rows)
    if len(registros) > limite:
        out.append(f"... {len(registros) - limite} resultado(s) adicionais.")
    return "\n".join(out)
